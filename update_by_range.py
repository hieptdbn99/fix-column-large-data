"""
update_by_range.py
===================
Updates CSV and XLSX files within a specific date range.
Usage: python update_by_range.py <start_date> <end_date>
Example: python update_by_range.py 21-03-2026 23-03-2026
"""

import os
import sys
import logging
import time
from datetime import datetime, timedelta
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import openpyxl

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
BASE_FOLDER = r"E:\mesv3"
LOG_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "range_update.log")

CSV_TARGET_COLUMNS = [
    "BucCoverQR", "BacketBarCode", "BendingDistanceValue", "PressureTime",
    "Temp 1", "Temp 2", "Temp 3", "Temp 4",
    "U1", "U2", "U3", "L1", "L2", "L3", "D1", "D2", "D3", "R1", "R2", "R3",
    "Result", "Remark", "Date", "Time",
]

CSV_RENAME_MAP = {
    "temp1":   "Temp 1", "temp2":   "Temp 2", "temp3":   "Temp 3", "temp4":   "Temp 4",
    "results": "Result",
}

# ─────────────────────────────────────────────
# CORE LOGIC (Integrated)
# ─────────────────────────────────────────────

def _normalise_csv_header(raw_name: str) -> str:
    stripped = raw_name.strip()
    return CSV_RENAME_MAP.get(stripped.lower(), stripped)

def process_csv_file(file_path_path: Path) -> tuple[bool, str | None]:
    try:
        with open(file_path_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        if not lines: return False, None
        header_line = lines[0].rstrip("\r\n")
        raw_fields  = header_line.split(",")
        norm_fields = [_normalise_csv_header(c) for c in raw_fields]
        if norm_fields == CSV_TARGET_COLUMNS: return False, None
        col_index = {name: i for i, name in enumerate(norm_fields)}
        missing = set(CSV_TARGET_COLUMNS) - set(col_index)
        if missing - {"Remark"}:
            return False, f"SKIP CSV missing cols {missing - {'Remark'}}: {file_path_path.name}"
        new_lines = [",".join(CSV_TARGET_COLUMNS) + "\n"]
        for line in lines[1:]:
            values = line.rstrip("\r\n").split(",")
            while len(values) < len(raw_fields): values.append("")
            new_values = []
            for col_name in CSV_TARGET_COLUMNS:
                idx = col_index.get(col_name)
                new_values.append(values[idx] if idx is not None and idx < len(values) else "")
            new_lines.append(",".join(new_values) + "\n")
        with open(file_path_path, "w", encoding="utf-8", newline="") as f:
            f.writelines(new_lines)
        return True, None
    except Exception as exc: return False, f"ERROR CSV {file_path_path.name}: {exc}"

def process_xlsx_file(file_path_path: Path) -> tuple[bool, str | None]:
    try:
        wb = openpyxl.load_workbook(file_path_path)
        sheet = wb.active
        headers = [cell.value for cell in sheet[1]]
        idx_result, idx_remark = -1, -1
        for i, h in enumerate(headers):
            if h is None: continue
            h_str = str(h).strip().lower()
            if h_str == "result": idx_result = i
            elif h_str == "remark": idx_remark = i
        if idx_result == -1 or idx_remark == -1:
            return False, f"SKIP XLSX: 'Result' or 'Remark' not found in {file_path_path.name}"
        if idx_result == idx_remark - 1: return False, None
        all_data = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            r_list = list(row)
            v_res = r_list.pop(idx_result)
            new_idx = idx_remark if idx_result > idx_remark else idx_remark - 1
            r_list.insert(new_idx, v_res)
            all_data.append(r_list)
        sheet.delete_rows(1, sheet.max_row)
        for r_idx, r_data in enumerate(all_data, 1):
            for c_idx, val in enumerate(r_data, 1):
                sheet.cell(row=r_idx, column=c_idx, value=val)
        wb.save(file_path_path)
        return True, None
    except Exception as exc: return False, f"ERROR XLSX {file_path_path.name}: {exc}"

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print("Usage: python update_by_range.py DD-MM-YYYY DD-MM-YYYY")
        print("Example: python update_by_range.py 21-03-2026 23-03-2026")
        return

    try:
        start_date = datetime.strptime(sys.argv[1], "%d-%m-%Y")
        end_date   = datetime.strptime(sys.argv[2], "%d-%m-%Y")
    except ValueError as e:
        print(f"Error parsing dates: {e}")
        return

    if start_date > end_date:
        print("Error: Start date must be before or equal to end date.")
        return

    logger = logging.getLogger("range_update")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    logger.info("=" * 60)
    logger.info(f"Date Range Update: {sys.argv[1]} to {sys.argv[2]}")
    logger.info("=" * 60)

    files_to_process = []
    current_date = start_date
    while current_date <= end_date:
        date_str = current_date.strftime("%Y-%m-%d")
        # Structure: E:\mesv3\YYYY\MM\DD
        path_parts = [
            BASE_FOLDER,
            current_date.strftime("%Y"),
            current_date.strftime("%m"),
            current_date.strftime("%d")
        ]
        folder_path = os.path.join(*path_parts)
        
        if os.path.isdir(folder_path):
            for f in os.listdir(folder_path):
                f_path = os.path.join(folder_path, f)
                if not os.path.isfile(f_path): continue
                
                f_upper = f.upper()
                if f_upper.startswith("VNA") and f_upper.endswith(".CSV"):
                    files_to_process.append(f_path)
                elif f == f"{date_str}.xlsx":
                    files_to_process.append(f_path)
        
        current_date += timedelta(days=1)

    if not files_to_process:
        logger.warning("No matching files found in the specified range.")
        return

    logger.info(f"Found {len(files_to_process)} files to process.")
    
    updated, skipped, errors = 0, 0, 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=8) as executor:
        futures = {}
        for f in files_to_process:
            p = Path(f)
            if p.suffix.lower() == ".csv":
                futures[executor.submit(process_csv_file, p)] = f
            elif p.suffix.lower() == ".xlsx":
                futures[executor.submit(process_xlsx_file, p)] = f
        
        for i, future in enumerate(as_completed(futures), 1):
            changed, msg = future.result()
            if msg:
                if "ERROR" in msg:
                    logger.error(msg)
                    errors += 1
                else:
                    logger.warning(msg)
                    skipped += 1
            elif changed: updated += 1
            else: skipped += 1
            
            if i % 10 == 0 or i == len(files_to_process):
                logger.info(f"Progress: {i}/{len(files_to_process)} | Updated={updated} | Skipped={skipped} | Errors={errors}")

    logger.info("=" * 60)
    logger.info(f"Done! Updated: {updated}, Skipped: {skipped}, Errors: {errors}")
    logger.info(f"Total time: {time.time() - start_time:.1f}s")
    logger.info(f"Log: {LOG_FILE}")

if __name__ == "__main__":
    main()

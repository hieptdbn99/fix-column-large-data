"""
run_update.py
==============
Unified script to:
1. Normalise VNA*.csv files (24 columns, specific order).
2. Reorder YYYY-MM-DD.xlsx files (full 24-column reorder: U,L,D,R).
"""

import os
import sys
import logging
import time
import re
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing
import openpyxl

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
BASE_FOLDER = r"E:\mesv4"
LOG_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "unified_update.log")

# CSV Target Columns
CSV_TARGET_COLUMNS = [
    "BucCoverQR", "BacketBarCode", "BendingDistanceValue", "PressureTime",
    "Temp 1", "Temp 2", "Temp 3", "Temp 4",
    "U1", "U2", "U3", "L1", "L2", "L3", "D1", "D2", "D3", "R1", "R2", "R3",
    "Result", "Remark", "Date", "Time",
]

CSV_RENAME_MAP = {
    "temp1":   "Temp 1", "temp2":   "Temp 2", "temp3":   "Temp 3", "temp4":   "Temp 4",
    "results": "Result",
}

# XLSX Target Headers (as they appear in Excel files, matching C# app excelmodel3)
XLSX_TARGET_HEADERS = [
    "BUC Cover\r\nQR코드", "Backet\r\nbar code", "압착 거리값", "압력 시간",
    "Temp 1", "Temp 2", "Temp 3", "Temp 4",
    "U1", "U2", "U3",
    "L1", "L2", "L3",
    "D1", "D2", "D3",
    "R1", "R2", "R3",
    "Result", "Remark",
    "생산일자", "생산시간",
]


def _normalise_xlsx_header(raw_value) -> str:
    if raw_value is None:
        return ""
    s = str(raw_value).strip()
    s = s.replace("_x000D_\n", "\r\n")
    return s


def _header_key(h: str) -> str:
    return h.replace("\r\n", "").replace("\r", "").replace("\n", "").replace(" ", "").lower()


# ─────────────────────────────────────────────
# CSV LOGIC
# ─────────────────────────────────────────────

def _normalise_csv_header(raw_name: str) -> str:
    stripped = raw_name.strip()
    return CSV_RENAME_MAP.get(stripped.lower(), stripped)

def process_csv_file(file_path_path: Path) -> tuple[bool, str | None]:
    try:
        with open(file_path_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        if not lines: return False, None

        header_line = lines[0].rstrip("\r\n")
        raw_fields  = header_line.split(",")
        norm_fields = [_normalise_csv_header(c) for c in raw_fields]

        if norm_fields == CSV_TARGET_COLUMNS: return False, None

        col_index = {name: i for i, name in enumerate(norm_fields)}
        missing = set(CSV_TARGET_COLUMNS) - set(col_index)
        if missing - {"Remark"}:
            return False, f"SKIP CSV missing cols {missing - {'Remark'}}: {file_path_path.name}"

        new_lines = [",".join(CSV_TARGET_COLUMNS) + "\n"]
        for line in lines[1:]:
            stripped = line.rstrip("\r\n")
            if not stripped:
                new_lines.append("\n")
                continue
            values = stripped.split(",")
            while len(values) < len(raw_fields): values.append("")
            new_values = []
            for col_name in CSV_TARGET_COLUMNS:
                idx = col_index.get(col_name)
                new_values.append(values[idx] if idx is not None and idx < len(values) else "")
            new_lines.append(",".join(new_values) + "\n")

        with open(file_path_path, "w", encoding="utf-8", newline="") as f:
            f.writelines(new_lines)
        return True, None
    except Exception as exc:
        return False, f"ERROR CSV {file_path_path.name}: {exc}"

# ─────────────────────────────────────────────
# XLSX LOGIC (full 24-column reorder)
# ─────────────────────────────────────────────

def process_xlsx_file(file_path_path: Path) -> tuple[bool, str | None]:
    try:
        wb = openpyxl.load_workbook(file_path_path)
        sheet = wb.active

        if sheet.max_row is None or sheet.max_row == 0:
            wb.close()
            return False, None

        raw_headers = [cell.value for cell in sheet[1]]
        norm_headers = [_normalise_xlsx_header(h) for h in raw_headers]

        col_index: dict[str, int] = {}
        for i, h in enumerate(norm_headers):
            key = _header_key(h)
            if key and key not in col_index:
                col_index[key] = i

        current_keys = [_header_key(h) for h in norm_headers]
        target_keys = [_header_key(h) for h in XLSX_TARGET_HEADERS]

        if current_keys == target_keys:
            wb.close()
            return False, None

        missing_keys = set(target_keys) - set(col_index.keys())
        remark_key = _header_key("Remark")
        unexpected_missing = missing_keys - {remark_key}
        if unexpected_missing:
            missing_names = []
            for mk in unexpected_missing:
                for th in XLSX_TARGET_HEADERS:
                    if _header_key(th) == mk:
                        missing_names.append(th)
                        break
            wb.close()
            return False, f"SKIP XLSX missing cols {missing_names}: {file_path_path.name}"

        all_rows = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            all_rows.append(list(row))

        new_rows = []
        for row_idx, row_data in enumerate(all_rows):
            while len(row_data) < len(raw_headers):
                row_data.append(None)
            new_row = []
            for target_h in XLSX_TARGET_HEADERS:
                tkey = _header_key(target_h)
                idx = col_index.get(tkey)
                if idx is not None and idx < len(row_data):
                    # Always use original data (including original header names)
                    new_row.append(row_data[idx])
                else:
                    # Missing column (e.g. Remark) - add target header name or empty
                    new_row.append(target_h if row_idx == 0 else "")
            new_rows.append(new_row)

        sheet.delete_rows(1, sheet.max_row)
        for r_idx, row_data in enumerate(new_rows, 1):
            for c_idx, value in enumerate(row_data, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        wb.save(file_path_path)
        return True, None

    except Exception as exc:
        return False, f"ERROR XLSX {file_path_path.name}: {exc}"

# ─────────────────────────────────────────────
# DISPATCHER
# ─────────────────────────────────────────────

def process_file_any(file_path_str: str) -> tuple[bool, str | None]:
    path = Path(file_path_str)
    ext = path.suffix.lower()
    if ext == ".csv":
        return process_csv_file(path)
    elif ext == ".xlsx":
        return process_xlsx_file(path)
    return False, f"Unknown extension: {path.name}"

# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    logger = logging.getLogger("unified")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")
    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)
    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    logger.info("=" * 60)
    logger.info("Unified Data Column Update (CSV + XLSX full reorder)")
    logger.info(f"Target Folder: {BASE_FOLDER}")
    logger.info("=" * 60)

    if not os.path.isdir(BASE_FOLDER):
        logger.error(f"Folder not found: {BASE_FOLDER}")
        return

    logger.info("Scanning for files...")
    xlsx_date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}(_\d+)?\.xlsx$")
    all_files = []

    for root, dirs, files in os.walk(BASE_FOLDER):
        for f in files:
            f_upper = f.upper()
            if f_upper.startswith("VNA") and f_upper.endswith(".CSV"):
                all_files.append(os.path.join(root, f))
            elif xlsx_date_pattern.match(f):
                all_files.append(os.path.join(root, f))

    total = len(all_files)
    if total == 0:
        logger.warning("No matching CSV or XLSX files found.")
        return

    logger.info(f"Found {total:,} total files to process.")
    num_workers = min(16, max(4, multiprocessing.cpu_count() * 2))

    updated, skipped, errors = 0, 0, 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = {executor.submit(process_file_any, f): f for f in all_files}
        for i, future in enumerate(as_completed(futures), 1):
            changed, msg = future.result()
            if msg:
                if "ERROR" in msg:
                    logger.error(msg)
                    errors += 1
                else:
                    logger.warning(msg)
                    skipped += 1
            elif changed:
                updated += 1
            else:
                skipped += 1

            if i % 50 == 0 or i == total:
                elapsed = time.time() - start_time
                logger.info(f"Progress: {i}/{total} | Updated={updated} | Skipped={skipped} | Errors={errors} | Time={elapsed:.1f}s")

    logger.info("=" * 60)
    logger.info(f"Done! Updated: {updated}, Skipped: {skipped}, Errors: {errors}")
    logger.info(f"Total time: {time.time() - start_time:.1f}s")
    logger.info(f"Log: {LOG_FILE}")

if __name__ == "__main__":
    main()

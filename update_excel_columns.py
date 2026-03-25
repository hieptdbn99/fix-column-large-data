"""
update_excel_columns.py
========================
Scan all YYYY-MM-DD.xlsx files in E:\\mesv3 and reorder columns to match
the standard 24-column structure (same order as C# app excelmodel3):

    BUC Cover QR코드, Backet bar code, 압착 거리값, 압력 시간,
    Temp 1, Temp 2, Temp 3, Temp 4,
    U1, U2, U3, L1, L2, L3, D1, D2, D3, R1, R2, R3,
    Result, Remark, 생산일자, 생산시간

Known source variations:
  - Missing 'Remark' column (2025/01) -> added with empty value
  - Column order L,R,U,D instead of U,L,D,R (2025/05+) -> reordered
  - Result after Remark -> moved before Remark
"""

import os
import sys
import logging
import time
import re
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing
import openpyxl

# ─────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────
BASE_FOLDER = r"E:\mesv4"
LOG_FILE    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "update_excel_order.log")

# The exact target column order (24 columns) - header names as they appear in XLSX
# This matches what the C# app (excelmodel3) writes
XLSX_TARGET_HEADERS = [
    "BUC Cover\r\nQR코드", "Backet\r\nbar code", "압착 거리값", "압력 시간",
    "Temp 1", "Temp 2", "Temp 3", "Temp 4",
    "U1", "U2", "U3",
    "L1", "L2", "L3",
    "D1", "D2", "D3",
    "R1", "R2", "R3",
    "Result", "Remark",
    "생산일자", "생산시간",
]


def _normalise_xlsx_header(raw_value) -> str:
    """Normalise an XLSX header value to a comparable string."""
    if raw_value is None:
        return ""
    s = str(raw_value).strip()
    # openpyxl may store \\r\\n as _x000D_\\n - normalise
    s = s.replace("_x000D_\n", "\r\n")
    return s


def _header_key(h: str) -> str:
    """Create a lowercase key for matching, ignoring whitespace/newline variations."""
    return h.replace("\r\n", "").replace("\r", "").replace("\n", "").replace(" ", "").lower()


# ─────────────────────────────────────────────
# WORKER FUNCTION
# ─────────────────────────────────────────────

def process_xlsx_file(file_path_str: str) -> tuple[bool, str | None]:
    """
    Read an XLSX, reorder all 24 columns to match XLSX_TARGET_HEADERS.
    Missing 'Remark' column is added with empty values.
    Returns (changed: bool, error_or_warning: str | None).
    """
    file_path = Path(file_path_str)
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active

        if sheet.max_row is None or sheet.max_row == 0:
            wb.close()
            return False, None

        # Read raw headers from row 1
        raw_headers = [cell.value for cell in sheet[1]]
        norm_headers = [_normalise_xlsx_header(h) for h in raw_headers]

        # Build index: header_key -> column index (first occurrence)
        col_index: dict[str, int] = {}
        for i, h in enumerate(norm_headers):
            key = _header_key(h)
            if key and key not in col_index:
                col_index[key] = i

        # Check if already in correct order
        current_keys = [_header_key(h) for h in norm_headers]
        target_keys = [_header_key(h) for h in XLSX_TARGET_HEADERS]

        if current_keys == target_keys:
            wb.close()
            return False, None  # Already correct

        # Check which target columns are missing
        missing_keys = set(target_keys) - set(col_index.keys())
        remark_key = _header_key("Remark")

        # Allow missing Remark (will be filled with empty)
        unexpected_missing = missing_keys - {remark_key}
        if unexpected_missing:
            missing_names = []
            for mk in unexpected_missing:
                for th in XLSX_TARGET_HEADERS:
                    if _header_key(th) == mk:
                        missing_names.append(th)
                        break
            wb.close()
            return False, f"SKIP XLSX missing cols {missing_names}: {file_path.name}"

        # Read all data rows
        all_rows = []
        for row in sheet.iter_rows(min_row=1, values_only=True):
            all_rows.append(list(row))

        # Reorder all rows according to target order
        new_rows = []
        for row_idx, row_data in enumerate(all_rows):
            # Pad row if shorter than headers
            while len(row_data) < len(raw_headers):
                row_data.append(None)

            new_row = []
            for target_h in XLSX_TARGET_HEADERS:
                tkey = _header_key(target_h)
                idx = col_index.get(tkey)
                if idx is not None and idx < len(row_data):
                    # Always use original data (including original header names)
                    new_row.append(row_data[idx])
                else:
                    # Missing column (e.g. Remark) - add target header name or empty
                    if row_idx == 0:
                        new_row.append(target_h)
                    else:
                        new_row.append("")
            new_rows.append(new_row)

        # Clear sheet and write reordered data
        sheet.delete_rows(1, sheet.max_row)
        for r_idx, row_data in enumerate(new_rows, 1):
            for c_idx, value in enumerate(row_data, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value)

        wb.save(file_path)
        return True, None

    except Exception as exc:
        return False, f"ERROR {file_path.name}: {exc}"


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────

def main():
    logger = logging.getLogger("excel_main")
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s")

    fh = logging.FileHandler(LOG_FILE, encoding="utf-8")
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    logger.info("=" * 60)
    logger.info("Excel Column Order Update (Full 24-column reorder: U,L,D,R)")
    logger.info(f"Base folder : {BASE_FOLDER}")
    logger.info("=" * 60)

    if not os.path.isdir(BASE_FOLDER):
        logger.error(f"Folder not found: {BASE_FOLDER}")
        return

    logger.info("Scanning for YYYY-MM-DD*.xlsx files...")
    date_pattern = re.compile(r"^\d{4}-\d{2}-\d{2}(_\d+)?\.xlsx$")

    xlsx_files = []
    for root, dirs, files in os.walk(BASE_FOLDER):
        for f in files:
            if date_pattern.match(f):
                xlsx_files.append(os.path.join(root, f))

    total = len(xlsx_files)
    if total == 0:
        logger.warning("No matching YYYY-MM-DD.xlsx files found.")
        return

    logger.info(f"Found {total:,} Excel file(s).")

    num_workers = min(8, max(2, multiprocessing.cpu_count()))
    logger.info(f"Starting processing with {num_workers} workers...")

    updated = 0
    skipped = 0
    errors  = 0
    start_time = time.time()

    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        future_to_path = {executor.submit(process_xlsx_file, f): f for f in xlsx_files}

        for i, future in enumerate(as_completed(future_to_path), 1):
            changed, msg = future.result()
            if msg:
                if msg.startswith("ERROR"):
                    logger.error(msg)
                    errors += 1
                else:
                    logger.warning(msg)
                    skipped += 1
            elif changed:
                updated += 1
            else:
                skipped += 1

            if i % 10 == 0 or i == total:
                elapsed = time.time() - start_time
                logger.info(f"Progress: {i}/{total} | Updated={updated} | Skipped={skipped} | Errors={errors} | Time={elapsed:.1f}s")

    logger.info("=" * 60)
    logger.info(f"Finished. Total Updated: {updated}, Skipped: {skipped}, Errors: {errors}")
    logger.info(f"Total time: {time.time() - start_time:.1f}s")

if __name__ == "__main__":
    main()
